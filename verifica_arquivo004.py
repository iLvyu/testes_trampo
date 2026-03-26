# Script verifica arquivos e indica quais estão presentes, com destaque visual no Excel

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

caminho_excel = r"C:\fritando_neuronios\dev_mondial\teste_automacao\arquivos_teste\ambiente_leitura\ex_n_container.xlsx"
pasta_arquivos = r"R:\COMEX\MK BAHIA\REMOÇÃO\Minutas\2026"

coluna_nome = "Nr. Container"
coluna_status = "Status"

# ===== PASSO 1: LÊ E ATUALIZA O EXCEL =====
df = pd.read_excel(caminho_excel, engine="openpyxl")

arquivos_na_pasta = set()
for raiz, pastas, arquivos in os.walk(pasta_arquivos):
    for arquivo in arquivos:
        arquivos_na_pasta.add(arquivo)


def verifica(container):
    container = str(container).strip()
    for arquivo in arquivos_na_pasta:
        if container in arquivo:
            return "OK"
    return ""


df[coluna_status] = df[coluna_nome].apply(verifica)

df.to_excel(caminho_excel, index=False, engine="openpyxl")

# ===== PASSO 2: ABRE COM OPENPYXL E PINTA AS LINHAS NÃO ENCONTRADAS =====
wb = load_workbook(caminho_excel)
ws = wb.active

# Preenchimento amarelo
amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Descobre o número da coluna "Status"
coluna_status_idx = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == coluna_status:
        coluna_status_idx = col
        break

# Pinta as linhas onde Status está vazio
for row in range(2, ws.max_row + 1):
    status = ws.cell(row=row, column=coluna_status_idx).value
    if status != "OK":
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = amarelo

wb.save(caminho_excel)

print("Execução finalizada")
