import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ===== CONFIGURAÇÕES =====
caminho_excel = r"C:\fritando_neuronios\dev_mondial\teste_automacao\arquivos_teste\ambiente_leitura\ex_n_container.xlsx"
pasta_arquivos = r"R:\COMEX\MK BAHIA\REMOÇÃO\Minutas\2026"

col_container = "Nr. Container"
col_status = "Status"
col_arquivo = "Arquivo Encontrado"
col_link = "Link Arquivo"
# =========================

# ===== PASSO 1: LER EXCEL =====
df = pd.read_excel(caminho_excel, engine="openpyxl")

# Garante que as colunas existam
for col in [col_status, col_arquivo, col_link]:
    if col not in df.columns:
        df[col] = ""

# Lista arquivos com CAMINHO COMPLETO (inclusive subpastas)
arquivos = []
for raiz, pastas, arquivos_pasta in os.walk(pasta_arquivos):
    for arq in arquivos_pasta:
        arquivos.append(os.path.join(raiz, arq))


# ===== PASSO 2: VERIFICAÇÃO =====
def verifica(container):
    container = str(container).strip()
    for caminho in arquivos:
        if container in os.path.basename(caminho):
            return "OK", os.path.basename(caminho), caminho
    return "", "", ""


df[[col_status, col_arquivo, col_link]] = df[col_container].apply(
    lambda x: pd.Series(verifica(x))
)

df.to_excel(caminho_excel, index=False, engine="openpyxl")

# ===== PASSO 3: LINKS E CORES =====
wb = load_workbook(caminho_excel)
ws = wb.active

amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Mapeia índices das colunas
col_idx = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}

for row in range(2, ws.max_row + 1):
    status = ws.cell(row=row, column=col_idx[col_status]).value
    caminho = ws.cell(row=row, column=col_idx[col_link]).value

    if status == "OK" and caminho:
        # Link clicável que ABRE O ARQUIVO (qualquer extensão)
        ws.cell(row=row, column=col_idx[col_link]).hyperlink = f"file:///{caminho}"
        ws.cell(row=row, column=col_idx[col_link]).value = "Abrir arquivo"
    else:
        # Linha inteira em amarelo se não encontrado
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = amarelo

wb.save(caminho_excel)

print("Execução finalizada")
